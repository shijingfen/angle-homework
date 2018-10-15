import re
import requests
from  bs4 import BeautifulSoup
import xlrd
import datetime

file = xlrd.open_workbook(r'C:\Users\fariyland\Desktop\a.xlsx')
tables = file.sheets()
table = file.sheets()[0] #获取第一个sheet#
rows = table.nrows       #行数
cols = table.ncols       #列数
tds=[]
i=1
data = []

while i<= rows-1:
    row = table.row_values(i) #获取第i行数据#
    col = table.col_values(0)#获取第1列数据#
    value = table.row_values(i)[0]    #获取第i行，第1列的数据#
    cell = table.cell(i, 0)       #读取第i行，第1列的单元格，这个方法获取的是单元格，并不是单元格中的值
    cell_value = table.cell(i, 0).value  #单元格中的值
    tds.append(cell_value)
    i+=1
cookie={
        "ASP.NET_SessionId":"511wpqgt0aouio3cytulesqe",
         ".ASPXUSERDEMO":"084CE0E4BA4A4A9B745B41F9AC4CDCA603F7F12E31D9FDA33EF0AF713E9B530D47CE0966D79D56CDBC8F0CFBF8E5E94D6152DC26EDA493A15D0E49E05BA7F06F348C926CC2E712918B9BC1802ADE40F999398077D47F5582368E6050DDF928DC3B6B79F5FD191C5B1CA12A480859B2D8B666A30FF083AED9D991865304CBE8ADA63E701FA7F076F83966F045F7572D3F"
        }
s=requests.Session()
for td in tds:
        url = ("http://172.17.237.76/serviceplatform/IPOC/Change/ChangeDetail.aspx?Id="+str(td))
        # print(url)
        from_data={"id":"14950"}
        rev=s.post(url,cookies=cookie,data=from_data)
        # print(rev.text)
        soup = BeautifulSoup(rev.text,'html.parser')
        PlanedTestTime = soup.select('span#lblPlanedTestTime')
        PTT=re.match(r".*[\>](.*)[\<].*",str(PlanedTestTime))#计划测试时间
        PlanedIssueTime = soup.select('span#lblPlanedIssueTime')
        PIT = re.match(r".*[\>](.*)[\<].*",str(PlanedIssueTime))#计划发布时间
        PlanedOnlineTime = soup.select('span#lblPlanedOnlineTime')
        PLT = re.match(r".*[\>](.*)[\<].*",str(PlanedOnlineTime))#计划上线时间#
        pattern = re.compile("已选模块.*")
        td_cvs_data = soup.find('td',text=pattern)
        print (td_cvs_data)
        td_cvs = td_cvs_data.next_sibling.next_sibling.string.strip()
        # print (td_cvs)
        
        PlanedTestTime = PTT.groups(1)[0]
        PlanedIssueTime = PIT.groups(1)[0]
        PlanedOnlineTime = PLT.groups(1)[0]
        data.append((td,td_cvs,PlanedTestTime,PlanedIssueTime,PlanedOnlineTime))   #构建输出数据
        print("td===>",td)
        print("CVS===>",td_cvs)
        print("PlanedTestTime===>",PlanedTestTime)
        print("PlanedIssueTime==>",PlanedIssueTime)
        print("PlanedOnlineTime==>",PlanedOnlineTime)

# print(data)

'''输出文件'''
from openpyxl import Workbook
wb = Workbook()
sheet = wb.active
sheet.title = "output"
sheet["A1"] = "TD"
sheet["B1"] = "CVS"
sheet["C1"] = "计划测试时间"
sheet["D1"] = "计划发布时间"
sheet["E1"] = "计划上线时间"


for i in range(len(data)):
    sheet["A%d"%(i+2)] = data[i][0]
    sheet["B%d" % (i + 2)] = data[i][1]
    sheet["C%d" % (i + 2)] = data[i][2]
    sheet["D%d" % (i + 2)] = data[i][3]
    sheet["E%d" % (i + 2)] = data[i][4]
wb.save('C:\\Users\\fariyland\\Desktop\\b.xlsx')







